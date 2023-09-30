
from openpyxl import Workbook, load_workbook
from os import path

import logging
logger = logging.getLogger(__name__)


def _load_database(db_path=None):
    if path.exists(db_path):
        wb = load_workbook(filename=db_path)
        return wb
    else:
        wb = Workbook()
        wb.save(filename=db_path)
        return wb


def make_database(db_path=None) -> bool:
    """Makes a new database in db_path

    Parameters
    ----------
    db_path : str
        File path of called database

    Returns
    -------
    bool
    """
    result = _load_database(db_path)
    if result:
        return True
    return result


def get_tables(db_path=None) -> list:
    """Get all tables in db_path

    Parameters
    ----------
    db_path : str
        File path of called database

    Returns
    -------
    list
    """
    wb = _load_database(db_path)
    return wb.sheetnames


def add_table(db_path=None, table_name='', **kwargs) -> bool:
    """Add new table to db_path

    Parameters
    ----------
    db_path : str
        File path of called database
    table_name : str
        Name of new table

    Returns
    -------
    bool
    """
    wb = _load_database(db_path)
    ws = wb.create_sheet("NewSheet")
    ws.title = table_name
    wb.save(filename=db_path)
    return True


def get_fields(db_path=None, table_name='') -> list:
    """Gets all fields from table_name

    Parameters
    ----------
    db_path : str
        File path of called database
    table_name : str
        Table name to search for fields

    Returns
    -------
    list
    """
    wb = _load_database(db_path)
    ws = wb[table_name]
    try:
        result = [(str(c.value), 'VARCHAR') for c in list(ws.rows)[0]]
        return result
    except IndexError as e:
        logger.error(e, exc_info=True)
        return []


def add_field(db_path=None, table_name='', field_info='') -> bool:
    """Add new field to table

    Parameters
    ----------
    db_path : str
        File path of called database
    table_name : str
        Table name to insert new field
    field_info : str
        Field information

    Returns
    -------
    bool
    """
    wb = _load_database(db_path)
    ws = wb[table_name]
    cols = [i for i in ws.iter_cols()]
    ws.cell(row=1, column=len(cols)+1, value=field_info)
    wb.save(filename=db_path)
    return True


def get_samples(db_path=None, table_name='', field_name='*', search_query='*') -> list:
    """Get samples of specified table_name and field_name

    Parameters
    ----------
    db_path : str
        File path of called database
    table_name : str
        Table name to insert new field
    field_name : str
        Name of field(s) to search
    search_query : str
        Search query, default '*'

    Returns
    -------
    list
    """
    wb = _load_database(db_path)
    ws = wb[table_name]
    result = []
    for row in list(ws.rows)[1::]:
        result.append([c.value for c in row])
    if search_query != '*':
        results = []
        for r in result:
            status = False
            for d in r:
                if str(search_query) in str(d):
                    status = True
            if status:
                results.append(r)
    else:
        results = result
    return results


def add_sample(db_path=None, table_name='', fields=None, data=None) -> bool:
    """Add sample to specified table_name and field_name

    Parameters
    ----------
    db_path : str
        File path of called database
    table_name : str
        Table name to insert new field
    fields : list
        Name of field(s)
    data : list
        List of data to add

    Returns
    -------
    bool
    """
    wb = _load_database(db_path)
    ws = wb[table_name]
    rows = [i for i in ws.iter_rows()]
    for c, d in enumerate(data):
        ws.cell(row=len(rows)+1, column=c+1, value=d)
    wb.save(filename=db_path)
    return True
